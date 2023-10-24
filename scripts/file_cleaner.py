import os
import re
import locale
import datetime
import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.utils.cell import coordinate_to_tuple


class FileCleaner:
    def __init__(self, main=False):
        self.qlik_file_name = 'data_for_qlik.parquet'
        self.current_dir = '..' if main else '.'
        self.unidades = pd.read_csv(f'{self.current_dir}/data/static/conversion_unidades.csv')
        self.catalogo = pd.read_parquet(f'{self.current_dir}/data/static/catalogo_productos.parquet')

    def __limpia_cadena(self, cadena):
        # Quitamos tabulaciones
        aux = str(cadena).replace('\t', '')
        aux = aux.rstrip()
        aux = aux.replace('´', '')
        aux = re.sub(r'\.\d+', '', aux)
        aux = re.sub(r'-\d+', '', aux)
        aux = re.sub(r'^ANSA', '', aux)
        # Tratamos de convertir a int, y luego lo regresamos a string
        try:
            aux = str(int(aux))
        except:
            aux = str(aux)
        return aux
    
    def __open_qlik_file(self):
        tipo_de_archivo = self.qlik_file_name.split('.')[-1]
        file_to_open = f'{self.current_dir}/data/{self.qlik_file_name}'
        # Verificar que el archivo verdaderamente exista, o crear uno nuevo en caso contrario
        if self.qlik_file_name not in os.listdir(f'{self.current_dir}/data'):
            # Creamos el archivo dummy solo para que no truene el resto de cosas
            return pd.DataFrame(columns=['sku', 'planeado', 'producido', 'tipo', 'planta', 'inicio_semana',
                    'inicio_semana_real', 'semana', 'anio', 'porcentaje', 'completado',
                    'superado', 'inferior', 'terminado', 'estatus', 'familia', 'marca',
                    'descripcion'])
        if tipo_de_archivo == "csv":
            data = pd.read_csv(file_to_open)
        elif tipo_de_archivo == "parquet":
            data = pd.read_parquet(file_to_open)
        return data

    def __asigna_status(self, x):
        if x == 1:
            return "COMPLETADO"
        if x < 1:
            return "PENDIENTE"
        else:
            return "SUPERADO"

    def __find_value(self, hoja, col, initial_row=1, text='COD', max_consecutive_nones=10):
        """
        hoja: la hoja sobre la que se está trabajando
        col: número de la columna sobre la que se desea iterar
        initial_row: la primera fila a partir de la cual se desea empezar a buscar
        text: el valor específico que se busca en la columna
        max_consecutive_nones: número máximo de nones antes de parar
        """
        # Si nos dan la letra de la columna la convertimos a número para poder trabajar
        if type(col) is str:
            col = coordinate_to_tuple(f'{col}1')[1]
            
        found_cells = []
        row = initial_row
        consecutive_nones = 0
        while consecutive_nones < max_consecutive_nones:
            value = hoja.cell(row, col).value
            # Hay dos opciones: value es el valor que buscamos o no
            if value == text:
                found_cells.append(row)
            else: # si no es el valor que buscamos tenemos que ver si es None o no
                if value is None: # Si el valor es None, tenemos que agregar 1 al contador de Nones
                    consecutive_nones += 1
                else: # Si el valor no es None, regresamos a 0 el contador de consecutive Nones
                    consecutive_nones = 0
            row += 1
        return found_cells
    
    def __obtener_primer_dia_semana(self, numero_semana, anio):
        numero_semana, anio = int(numero_semana), int(anio)
        # Crear un objeto de fecha y tiempo para el primer día del año dado
        primer_dia_anio = datetime.datetime(anio, 1, 1)

        # Calcular la diferencia en días entre el primer día del año y el primer día de la semana dada
        dias_diferencia = (numero_semana - 1) * 7

        # Obtener el primer día de la semana sumando la diferencia de días al primer día del año
        primer_dia_semana = primer_dia_anio + datetime.timedelta(days=dias_diferencia)

        return primer_dia_semana
    
    def __extraer_numeros(self, texto):
        patron = r'\d+'
        numeros = re.findall(patron, texto)
        return numeros
    
    def __extraer_nombres_meses(self, texto):
        meses_espanol = {
            'ene': 'enero',
            'feb': 'febrero',
            'mar': 'marzo',
            'abr': 'abril',
            'may': 'mayo',
            'jun': 'junio',
            'jul': 'julio',
            'ago': 'agosto',
            'sep': 'septiembre',
            'oct': 'octubre',
            'nov': 'noviembre',
            'dic': 'diciembre'
        }
        patron = r'\b(?:ENERO|ENE|FEBRERO|FEB|MARZO|MAR|ABRIL|ABR|MAYO|MAY|JUNIO|JUN|JULIO|JUL|AGOSTO|AGO|SEPTIEMBRE|SEP|OCTUBRE|OCT|NOVIEMBRE|NOV|DICIEMBRE|DIC)\b'
        nombres_meses = re.findall(patron, texto, re.IGNORECASE)
        nombres_meses = [meses_espanol[nombre_mes.lower()] if len(nombre_mes) == 3 else nombre_mes for nombre_mes in nombres_meses ]
        return nombres_meses
    
    def __crear_fecha(self, numero_dia, nombre_mes, year=None):
        fecha_actual = datetime.datetime.now()
        if year is None:
            año_actual = fecha_actual.year
        else:
            año_actual = year
        
        # Establecer el idioma en español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

        
        fecha_str = f"{numero_dia} {nombre_mes} {año_actual}"
        fecha = datetime.datetime.strptime(fecha_str, "%d %B %Y")
        
        # Restaurar la configuración regional predeterminada
        locale.setlocale(locale.LC_TIME, '')
        
        return fecha
    
    def __generate_date(self, nombre_archivo, year=None):
        '''
        Esta función recibe una cadena de texto que contiene al menos un número y un nombre de mes.
        Extrae los número y nombres de mex que existan ahí y los convierte a una datetime.
        En caso de que no se especifique el año, se toma el año actual.
        '''
        dia = self.__extraer_numeros(nombre_archivo)[0]
        mes = self.__extraer_nombres_meses(nombre_archivo)[0]
        try:
            fecha = self.__crear_fecha(dia, mes, year)
        except:
            print(f'year was not successful in {nombre_archivo}')
        return fecha
    
    def __obtener_columnas_en_fila(self, hoja, numero_fila, num_nones_consecutivos):
        diccionario_columnas = {}
        nones_consecutivos = 0
    
        for columna, celda in enumerate(hoja[numero_fila], start=1):
            valor_celda = celda.value
            
            if valor_celda is not None:
                diccionario_columnas[valor_celda] = columna
                nones_consecutivos = 0
            else:
                nones_consecutivos += 1
                if nones_consecutivos >= num_nones_consecutivos:
                    break
        
        return diccionario_columnas

    def __actualiza_catalogo(self, new_data:pd.DataFrame):
        new_catalogo = (
            new_data
            [['sku', 'descripcion']]
            .merge(self.catalogo, on='sku', how='outer')
            .fillna(value={'familia':'EXTERNO', 'marca':'BAYER'})
            .assign(descripcion=lambda df: df.descripcion_y.fillna(df.descripcion_x).apply(self.__limpia_descripcion))
            .drop(['descripcion_x', 'descripcion_y'], axis=1)
            # Vamos a agrupar para cada sku, contar cuántas descripciones distintas hay
            .drop_duplicates()
            .groupby(['sku', 'familia', 'marca'])
            .agg(descripciones=pd.NamedAgg('descripcion', set))
            .assign(descripcion=lambda x: x.descripciones.apply(list).apply(self.__selecciona_descripcion))
            .drop(['descripciones'], axis=1)
            .reset_index()
            .query('sku != "0"')
        )
        new_catalogo.to_parquet(f'{self.current_dir}/data/static/catalogo_productos.parquet', index=False)
        self.catalogo = new_catalogo

    def __limpia_descripcion(self, cadena:str):
        if cadena is None:
            return 'SIN DESCRIPCION'
        cadena = cadena.replace(',', ' ')
        cadena = cadena.strip()
        return cadena
    
    def __selecciona_descripcion(self, lista_descripciones:list):
        '''
        Esta función recibe una lista con todas las descripciones
        que tiene un producto dado y elige la mejor (la más corta)
        Si en alguna hay una descripción del tipo "SIN DESCRIPCION"
        se omite en caso de que haya otras
        '''
        if len(lista_descripciones) == 1:
            return lista_descripciones[0]
        if "SIN DESCRIPCION" in lista_descripciones:
            lista_descripciones.remove("SIN DESCRIPCION")
        return min(lista_descripciones, key=len)

    def __save_updated_df(self, updated_data: pd.DataFrame):
        # Ahora guardamos los datos
        # Abrimos el archivo histórico, le pegamos los datos y los volvemos a guardar
        previous_data = self.__open_qlik_file()
        concated_data = pd.concat([previous_data, updated_data],ignore_index=True)
        sin_duplicados = concated_data.drop_duplicates(subset=['sku', 'semana', 'anio'], keep='last')
        # Actualizamos el catalogo, para asegurarnos de tener los más nuevos
        self.__actualiza_catalogo(sin_duplicados)
        # Tengo que asegurarme de que los datos de qlik estén limpios, porque siguen sucios
        # Ya sabemos que todos los productos aquí tienen alguna descripción disponible, entonces es hacer el match y ya
        sin_duplicados = (
            sin_duplicados
            .drop(['descripcion', 'familia', 'marca'], axis=1)
            .merge(self.catalogo, on='sku', how='left')
        )
        print(sin_duplicados.query('sku == "ABC1X20"'))
        # Guardamos el archivo
        tipo_de_archivo = self.qlik_file_name.split('.')[-1]
        if tipo_de_archivo == "csv":
            sin_duplicados.to_csv(f'{self.current_dir}/data/{self.qlik_file_name}', index=False)
        elif tipo_de_archivo == "parquet":
            sin_duplicados.to_parquet(f'{self.current_dir}/data/{self.qlik_file_name}', index=False)

    def __filtrar_valores_numericos(self, lista):
        valores_numericos = []
        for valor in lista:
            try:
                numero = float(valor)
                valores_numericos.append(valor)
            except ValueError:
                pass
        return valores_numericos
    
    def __extrae_datos_columnas(self, working_sheet, starting_row, dict_cols, ending_row=None, cols_to_extract=None, output_columns=None, rows_to_extract=100):
        '''
        Esta función extrae los datos de una hoja de excel y lo regresa como un dataframe.
        working_sheet: la hoja de excel
        starting_row: el número de la fila en la que empiezan los valores
        dict_cols: un diccionario de la siguiente forma {'col_name_1':numero_columna_1, 'col_name_2':numero_columna_2}
        ending_row: el número de la final donde terminan los valores (si es None, se extraen rows_to_extract filas tras starting_row)
        cols_to_extract: una lista con los nombres de las columnas que se desean seleccionar (None->todas las dict_cols)
        output_columns: los nombres de las columnas que se usarán en el DataFrame (None-> las mismas que dict_cols)
        rows_to_extract: el máximo número de filas a extraer. Solo se considera si ending_row es None
        '''
        # Definimos los parámetros necesarios
        
        # LAST_ROW
        if ending_row is None:
            last_row = starting_row + rows_to_extract
        else:
            last_row = ending_row
            
        # COLS_TO_EXTRACT
        if cols_to_extract is None:
            cols_to_extract = list(dict_cols.keys())
        
        # OUTPUT_COLUMNS
        if output_columns is None:
            output_columns = cols_to_extract
            
        listas_datos = [[] for _ in range(len(cols_to_extract))]
        for row in np.arange(starting_row, last_row):
            for lista, col_name in zip(listas_datos, cols_to_extract):
                lista.append(working_sheet.cell(row, dict_cols[col_name]).value)
                
        df = pd.DataFrame(np.array(listas_datos).transpose(), columns=output_columns)
        df.dropna(how='all', inplace=True)
        return df  
    
    def __find_last_data_row(self, ws, column_num, max_consecutive_nones, start_row):
        consecutive_nones = 0
        last_data_row = start_row - 1

        for row in ws.iter_rows(min_row=start_row, min_col=column_num, max_col=column_num):
            cell_value = row[0].value
            if cell_value is None:
                consecutive_nones += 1
                if consecutive_nones >= max_consecutive_nones:
                    return last_data_row
            else:
                last_data_row = row[0].row
                consecutive_nones = 0

        return last_data_row
    
    def __get_previous_processed_weeks(self, planta: str, tipo: str):
        previous_data = self.__open_qlik_file()
        return list(previous_data.query('planta == @planta and tipo == @tipo').semana.unique())

    def __clean_data_frame(self, df_to_clean, fecha_formateada, year, semana):
        aux = (df_to_clean
        .query('planeado != 0')
        .dropna(subset=['sku', 'planeado'])
        .assign(
            sku=lambda x: x.sku.apply(self.__limpia_cadena),
            tipo='TBD',
            planta='LERMA',
            inicio_semana=fecha_formateada,
            inicio_semana_real=fecha_formateada,
            anio=year,
            semana=semana,
            planeado=lambda df: df.apply(lambda x: pd.to_numeric(x.planeado, errors='coerce'), axis=1),
            producido=lambda df: df.apply(lambda x: pd.to_numeric(x.producido, errors='coerce'), axis=1),
            # Calculamos el porcentaje de producción
            porcentaje=lambda x: np.round(x.producido.astype(float)/x.planeado.astype(float), 2),
            # Agregamos indicadoras del estatus del producto
            completado=lambda x: (x.porcentaje == 1).astype('int8'),
            superado=lambda x: (x.porcentaje > 1).astype('int8'),
            inferior=lambda x: (x.porcentaje < 1).astype('int8'),
            terminado=lambda x: (x.porcentaje >= 1).astype('int8'),
            estatus=lambda df: df.porcentaje.apply(self.__asigna_status),
            descripcion=lambda df: df.descripcion.replace('#N/A', None)
        )
        .merge(self.catalogo, on='sku', how='left')
        .fillna(value={'porcentaje':0, 'familia':'EXTERNO', 'marca':'BAYER', 'descripcion':'EXTERNO'})
        .assign(descripcion=lambda df: df.descripcion_y.fillna(df.descripcion_x))
        .drop(['descripcion_x', 'descripcion_y'], axis=1)
        )
        return aux

    def __calculate_week_mode(self, date_list):
        date_list.sort()  # Ordenar las fechas
        start_date, end_date = date_list
        
        date_range = [start_date + datetime.timedelta(days=x) for x in range((end_date - start_date).days + 1)]
        

        week_numbers = np.array([date.isocalendar()[1] for date in date_range])  # Obtener números de semana como arreglo NumPy

        mode_week = np.argmax(np.bincount(week_numbers))  # Calcular la moda usando NumPy

        return mode_week, week_numbers
    
    def clean_liquidos(self, file):
        '''
        Esta función recibe dos posibles cosas: una ubicación de un archivo o un archivo de excel en bytes (file)
        Este archivo debe ser de líquidos y tener el formato esperado.
        '''

        # Definición de constantes
        nombre_hoja = 'Plan de Producción'
        celda_fecha = 'C5'
        celda_inicio_datos = 'B13'

        # Apertura del archivo
        archivo = xl.load_workbook(file, data_only=True)
        programa = archivo[nombre_hoja]

        coords_fecha = coordinate_to_tuple(celda_fecha)
        fechas = []
        for row in range(coords_fecha[0]-3, coords_fecha[0]+3):
            fechas.append(programa.cell(row, coords_fecha[1]).value)
        
        if all(x is None for x in fechas):
            print(f'Dates are not in the correct cell in {file.name}')
            return False
        fechas_aux = [aux for aux in fechas if aux is not None] # Aquí tengo dos fechas
        mode_week, all_weeks = self.__calculate_week_mode(fechas_aux)
        
        if all_weeks.size > 13:
            print(f"No se pudo entender la fecha en el archivo {file.name}")
            return False
        fecha = self.__obtener_primer_dia_semana(mode_week, datetime.datetime.now().year)
        
        #fecha = self.__obtener_primer_no_none(fechas)
        
        year = fechas_aux[0].year
        semana = fecha.isocalendar().week
        fecha_formateada = fechas_aux[0].strftime('%d/%m/%Y')
        inicio_semana_real = fecha.strftime('%d/%m/%Y')


        starting_data_row = self.__find_value(programa, celda_inicio_datos[0], text='CLAVE')[0]+1
        starting_data_col = coordinate_to_tuple(celda_inicio_datos)[1]

        # Ahora tomamos todos los datos de los skus
        # NOTA: se espera que los datos empiecen siempre en la celda B13
        # NOTA: no sé en qué medidas están. Supondré que son kg/lt
        skus = []
        descripciones = []
        programado = []
        fabricado = []
        for row in range(starting_data_row, starting_data_row + 250):
            skus.append(programa.cell(row, starting_data_col).value)
            descripciones.append(programa.cell(row, starting_data_col + 1).value)  # se espera que las descripciones estén a un lado de los ids
            programado.append(programa.cell(row, starting_data_col + 2).value) # se espera que lo programado esté a un lado de la descripción
            fabricado.append(programa.cell(row, starting_data_col + 4).value) # se espera que lo fabricado esté a dos lados de lo programado 

        # Creamos un dataframe para poder limpiar los datos
        df = (pd
         .DataFrame({'sku':skus, 'descripcion':descripciones, 'planeado':programado, 'producido':fabricado})
         .assign(
             planeado=lambda df: pd.to_numeric(df.planeado, errors='coerce'),
             producido=lambda df: pd.to_numeric(df.producido, errors='coerce')
         )
        )

        # Limpiamos los datos
        datos_limpios_aux = (df
         .dropna()
         .query('planeado != 0')
         .groupby(['sku', 'descripcion'])
         .agg(
             planeado=pd.NamedAgg('planeado', 'sum'),
             producido=pd.NamedAgg('producido', 'sum')
         )
         .reset_index()
         .assign(
             # Limpiamos los skus para eliminar caracteres no deseados
             sku=lambda df: df.sku.apply(self.__limpia_cadena),
             # Asignamos el tipo adecuado
             tipo='LIQUIDOS',
             planta='IZUCAR',
             # Asignamos los datos de la fecha
             inicio_semana=fecha_formateada,
             inicio_semana_real = inicio_semana_real,
             semana=semana,
             anio=year,
             # NOTA: puede hacer falta agregar la conversión a unidades adecuadas
             # Calculamos el porcentaje de producción
             porcentaje=lambda x: np.round(x.producido/x.planeado, 2),
             # Agregamos indicadoras del estatus del producto
             completado=lambda x: (x.porcentaje == 1).astype('int8'),
             superado=lambda x: (x.porcentaje > 1).astype('int8'),
             inferior=lambda x: (x.porcentaje < 1).astype('int8'),
             terminado=lambda x: (x.porcentaje >= 1).astype('int8'),
             estatus=lambda df: df.porcentaje.apply(self.__asigna_status)
         )
         .merge(self.catalogo, on='sku', how='left')
         .fillna(value={'porcentaje':0, 'familia':'EXTERNO', 'marca':'BAYER'})
         .assign(descripcion=lambda df: df.descripcion_y.fillna(df.descripcion_x))
         .drop(['descripcion_x', 'descripcion_y'], axis=1)
        )

        self.__save_updated_df(datos_limpios_aux)
        return True

    def clean_polvos(self, file):
        
        celda_fecha = "C5"
        # También hay que suponer que no hay fecha ahí, sino que está en el nombre del archivo
        # porque así es en la realidad. Sino se tiene que poner de forma manual
        celda_inicio_datos = "B10" # Celda donde se encuentra la primera clave

        row_inicio_datos = coordinate_to_tuple(celda_inicio_datos)[0]
        col_inicio_datos = coordinate_to_tuple(celda_inicio_datos)[1]

        # Abrimos el archivo y hoja
        wb = xl.load_workbook(file, data_only=True)
        programa = wb.active 
        
        # Encontramos la fila donde empiezan los datos y leemos las columnas
        starting_data_row = self.__find_value(programa, celda_inicio_datos[0], text='CLAVE', max_consecutive_nones=10)[0]+1 
        diccionario_cols = self.__obtener_columnas_en_fila(programa, starting_data_row-3, 15)
        
        # Generamos la fecha a partir del nombre del archivo
        # Vemos si hay alguna fecha en la celda esperada, sino la creamos nosotros
        if programa[celda_fecha].value is None:
            fecha = self.__generate_date(file.name, year=None)
        else:
            fecha = programa[celda_fecha].value
        year = fecha.year
        semana = fecha.isocalendar().week
        fecha_formateada = fecha.strftime('%d/%m/%Y')
        inicio_semana_real = self.__obtener_primer_dia_semana(semana, year).strftime('%d/%m/%Y')

        skus = []
        descripciones = []
        programado = []
        fabricado = []
        for row in range(row_inicio_datos, row_inicio_datos + 100):
            skus.append(programa.cell(row, col_inicio_datos).value)
            descripciones.append(programa.cell(row, col_inicio_datos + 1).value) 
            programado.append(programa.cell(row, diccionario_cols['KG PROGRAMADOS']).value)
            fabricado.append(programa.cell(row, diccionario_cols['KG FABRICADOS']).value)

        # Creamos un dataframe para poder limpiar los datos
        df = pd.DataFrame({'sku':skus, 'descripcion':descripciones, 'planeado':programado, 'producido':fabricado})

        # Limpiamos los datos
        limpio_aux = (df
         .dropna()
         .query('planeado != 0')
         .groupby(['sku', 'descripcion'])
         .agg(
             planeado=pd.NamedAgg('planeado', 'sum'),
             producido=pd.NamedAgg('producido', 'sum')
         )
         .reset_index()
         .assign(
             # Limpiamos los skus para eliminar caracteres no deseados
             sku=lambda df: df.sku.apply(self.__limpia_cadena),
             # Asignamos el tipo adecuado
             tipo='POLVOS',
             planta='IZUCAR',
             # Asignamos los datos de la fecha
             inicio_semana=fecha_formateada,
             inicio_semana_real = inicio_semana_real,
             semana=semana,
             anio=year,
             # NOTA: puede hacer falta agregar la conversión a unidades adecuadas
             # Calculamos el porcentaje de producción
             porcentaje=lambda x: np.round(x.producido.astype(float)/x.planeado.astype(float), 2),
             # Agregamos indicadoras del estatus del producto
             completado=lambda x: (x.porcentaje == 1).astype('int8'),
             superado=lambda x: (x.porcentaje > 1).astype('int8'),
             inferior=lambda x: (x.porcentaje < 1).astype('int8'),
             terminado=lambda x: (x.porcentaje >= 1).astype('int8'),
             estatus=lambda df: df.porcentaje.apply(self.__asigna_status)
         )
         .merge(self.catalogo, on='sku', how='left')
         .fillna(value={'porcentaje':0, 'familia':'EXTERNO', 'marca':'BAYER'})
         .assign(descripcion=lambda df: df.descripcion_y.fillna(df.descripcion_x))
         .drop(['descripcion_x', 'descripcion_y'], axis=1)
        )

        self.__save_updated_df(limpio_aux)

    def clean_lerma(self, file):
        wb = xl.load_workbook(file, data_only=True)

        # Vemos las semanas que ya habíamos procesado
        previous_processed_weeks = self.__get_previous_processed_weeks(planta='LERMA', tipo='TBD')

        sheets = self.__filtrar_valores_numericos(wb.sheetnames)
        clean_dfs = []
        processed_weeks = []

        for sheet in sheets:
            if int(sheet) in previous_processed_weeks:
                continue
            processed_weeks.append(int(sheet))
            ws = wb[sheet]
            # Creamos la fecha a partir del nombre de la hoja
            # En este caso se extrae del nombre de la hoja
            fecha = self.__obtener_primer_dia_semana(int(sheet), datetime.datetime.now().year)
            fecha_formateada = fecha.strftime('%d/%m/%Y')
            year = fecha.year
            semana = int(sheet)

            # Primero encontramos en qué fila están los nombres de las columnas
            list_cols = self.__find_value(ws, col=1, text='Área') # Se supone que ahí esté eso, pero ya se les ocurrió cambiarlo
            if len(list_cols) == 0: # Si no encontró 'Área', entonces buscamos 'Equipo'
                list_cols = self.__find_value(ws, col=1, text='Equipo')
            if len(list_cols) == 0: # Si ninguno se encontró entonces ya al chile no hay nada y le decimos al usuario
                print(f'No se identificó el inicio de las filas con datos en la hoja {sheet}')
                continue
            row_cols = list_cols[0]
            
            # Encontramos los nombres de las columnas
            dict_cols = self.__obtener_columnas_en_fila(ws, row_cols, 10)
            
            # Encontramos la última fila con datos
            # Dado que tienen un desmadre de filas ocultas tenemos que considerar un número exagerado de consecutive nones
            last_row = self.__find_last_data_row(ws, dict_cols['Equipo'], 100, row_cols)


            # Tercero, extraemos los valores de las columnas correspondientes
            # Las columnas que queremos son las siguientes: Clave, Programado por semana y Producido
            cols_to_extract = ['Clave', 'Descripción', 'Programado por \nsemana', 'Producido']
            output_column_names = ['sku', 'descripcion', 'planeado', 'producido']
            df_datos = self.__extrae_datos_columnas(
                working_sheet=ws, 
                starting_row=row_cols+1, 
                dict_cols=dict_cols, 
                cols_to_extract=cols_to_extract, 
                output_columns=output_column_names
            )

            # Limpiamos los datos
            clean_df_sheet = self.__clean_data_frame(df_datos, fecha_formateada, year, semana)
            clean_dfs.append(clean_df_sheet) # Aquí están todas las semanas extraídas

        # Convertimos todos los dataframes de semanas extraídas a uno solo que es el que vamos a guardar
        if len(clean_dfs) > 0:
            extracted_data = pd.concat(clean_dfs, ignore_index=True)
            self.__save_updated_df(extracted_data)

        return processed_weeks

    def get_last_update_date(self):
        """
        Esta función regresa la última fecha actualizada de las diferentes cosas en un diccionario.
        El ususario puede tomar el que necesite.
        """
        datos = self.__open_qlik_file()

        fechas = (datos
         .assign(inicio_semana_real=lambda x: pd.to_datetime(x.inicio_semana_real, dayfirst=True))
         .groupby(['planta', 'tipo'])
         .agg(last_update=pd.NamedAgg('inicio_semana_real', 'max'))
        )

        dict_aux = (fechas
         .assign(last_update=lambda x: x.last_update.dt.strftime('%d-%m-%Y'))
         .reset_index()
         .set_index('tipo')
         .to_dict()
         ['last_update']
        )

        return dict_aux
    
    def get_cumplimiento_en_intervalo(self, inicio:pd.Timestamp, fin:pd.Timestamp) -> float:
        '''
        Esta función recibe dos timestamps, y encuentra el cumplimiento entre esas fechas.
        Input:
            inicio -> El domingo en el que empieza el intervalo
            fin -> El domingo en el que termina el intervalo (no incluido para el cálculo)

        Return: 
            float -> El porocentaje de cumplimiento en el intervalo dado

        Ejemplo:
            inicio = 13-agosto-2023
            fin = 20-agosto-2023
            El resultado será el porcentaje de cumplimiento en la semana correspondiente a las fechas
            14-agosto, 15-agosto, 16-agosto, 17-agosto, 18-agosto (porque el 13 es domingo, el 19 es sábado y el 20 de agosto no está incluido)
        '''
        datos_filtrados = self.filtra_datos_intervalo(inicio=inicio, fin=fin)
        # Ahora econtramos el porcentaje de cumplimiento en ese intervalo
        return datos_filtrados.terminado.mean()
    
    def get_total_products_on_interval(self, inicio:pd.Timestamp, fin:pd.Timestamp) -> (int, int, int):
        '''
        Esta función recibe un intervalo y calcula cuántos productos hay de cada tipo de producción, así como su total
        '''
        datos_filtrados = self.filtra_datos_intervalo(inicio=inicio, fin=fin)
        datos_agregados = (datos_filtrados
         .assign(sku=lambda df: df.sku.apply(self.__limpia_cadena))
         .groupby(['sku', 'familia'])
         .agg(total=pd.NamedAgg('tipo', 'count'))
         .reset_index()
        )
        
        total_prods = datos_agregados.sku.unique().size
        total_externos = datos_agregados.query('familia == "EXTERNO"').sku.unique().size
        total_dragon = total_prods - total_externos
        return total_prods, total_dragon, total_externos

    def filtra_datos_intervalo(self, inicio:pd.Timestamp, fin:pd.Timestamp) -> pd.DataFrame:
        '''
        Función que recibe dos timestamps y filtra los datos correspondientes.
        Regresa un dataframe
        '''
        datos = self.__open_qlik_file()
        
        inicio = pd.to_datetime(inicio.strftime('%m/%d/%Y'))
        fin= pd.to_datetime(fin.strftime('%m/%d/%Y'))
        datos = (datos
                 .assign(inicio_semana_real=lambda x: pd.to_datetime(x.inicio_semana_real, dayfirst=True))
                 .pipe(lambda df: df.loc[df.inicio_semana_real >= inicio])
                 .pipe(lambda df: df.loc[df.inicio_semana_real < fin])
                )
        
        return datos
    
    def get_products_on_interval(self, inicio:pd.Timestamp, fin:pd.Timestamp, dates_are_sundays=True) -> pd.DataFrame:
        '''
        Esta función recibe un tiempo de inicio, y uno de final, y devuelve el agregado de los productos que se hicieron 
        en dicho intervalo.
        '''
        if dates_are_sundays:
            display_date_inicio = inicio + datetime.timedelta(days=1)
            display_date_fin = fin - datetime.timedelta(days=2)
        else:
            display_date_inicio = inicio
            display_date_fin = fin

        grouping_columns = ['sku', 'descripcion', 'marca', 'familia', 'planta']
        datos_filtrados = self.filtra_datos_intervalo(inicio=inicio, fin=fin)
        datos_filtrados = (datos_filtrados
         .assign(sku=lambda df: df.sku.apply(self.__limpia_cadena))
         .groupby(grouping_columns)
         .agg(
             planeado=pd.NamedAgg('planeado', np.sum),
             producido=pd.NamedAgg('producido', np.sum),
             lista_datos=pd.NamedAgg('porcentaje', list)
         )
         .assign(
             inicio_intervalo=display_date_inicio.strftime('%d-%m-%Y'),
             fin_intervalo=display_date_fin.strftime('%d-%m-%Y')
         )
         .reset_index()
         .assign(cumplimiento=lambda x: np.round(x.producido/x.planeado, 2))
         .sort_values(by='cumplimiento')
         [['inicio_intervalo', 'fin_intervalo']+grouping_columns+['planeado', 'producido', 'cumplimiento', 'lista_datos']]
        )
        return datos_filtrados

    def get_column(self, column:str, unique:bool=True):
        '''
        Función que encuentra todas las familias disponibles
        '''
        datos = self.__open_qlik_file()
        assert column in datos.columns

        datos_columna = datos[column]
        if unique:
            datos_columna = datos_columna.unique()
            try:
                datos_columna.sort()
            except:
                datos_columna = [str(dato) for dato in datos_columna]
                datos_columna.sort()
        return datos_columna
    
    def get_filtered_values(self, columna:str, valor:str, return_column:str=None, unique:bool=True):
        '''
        Función que recibe un nombre de columna que debe estar en las columnas del df
        y un valor. Selecciona la columna y regresa los valores que estén en dicha columna y que 
        sean iguales al valor especificado.

        Si unique = True se regresan los valores únicos
        '''
        datos = self.__open_qlik_file()
        assert columna in datos.columns
        if return_column is None:
            return_column = columna
        else:
            assert return_column in datos.columns

        filtered_data = datos.query(f'{columna} == "{valor}"')[return_column]
        if unique:
            filtered_data = filtered_data.unique()
            filtered_data.sort()

        return filtered_data
    
    def get_values_for_sku(self, sku:str) -> pd.DataFrame:
        '''
        Esta función regresa todos los datos disponibles para un productos especificado
        '''
        datos = self.__open_qlik_file()
        orden_output = [
            'inicio_semana_real',
            'semana',
            'anio',
            'sku',
            'descripcion',
            'marca',
            'familia',
            'planta',
            'planeado',
            'producido',
            'porcentaje'
        ]
        return datos.query('sku == @sku').reset_index(drop=True)[orden_output].sort_values(['anio', 'semana'])
        

if __name__ == '__main__':
    print('file_cleaner is execting as main. This file is not intended for that use, problems might raise')

    fc = FileCleaner(main=True)
    print(fc.unidades)
    print(fc.catalogo)

    