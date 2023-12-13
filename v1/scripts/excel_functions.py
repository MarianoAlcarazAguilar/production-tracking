import pandas as pd
import numpy as np
import datetime
from openpyxl.utils.cell import coordinate_to_tuple
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class DateFunctions:
    def filtrar_valores_numericos(lista):
        if type(lista) is str:
            lista = lista.split()
        valores_numericos = []
        for valor in lista:
            try:
                numero = float(valor)
                valores_numericos.append(numero)
            except ValueError:
                pass
        return valores_numericos
    
    def asigna_num_mes(mes):
        '''
        Función que recibe un mes en formato String en español y lo convierte a número correspondiente.
        Básicamente es un diccionario de meses.
        '''
        mes = mes.lower()
        meses_aux = {'enero':1, 'febrero':2, 'marzo':3, 'abril':4, 'mayo':5, 'junio':6, 'julio':7, 'agosto':8,
            'septiembre':9, 'octubre':10, 'noviembre':11, 'diciembre':12, '':12, np.NaN:0, None:0,
            'ene':1, 'feb':2, 'mar':3, 'abr':4, 'may':5, 'jun':6, 'jul':7, 'ago':8, 'sep':9, 'oct':10,
            'nov':11, 'dic':12
        }
        return meses_aux[mes]
    
    def obtener_primer_dia_semana(numero_semana, anio):
        '''
        Esta función recibe un número de semana y un año. Resgresa un datetime
        con el primer día de la semana especificada.
        '''
        numero_semana, anio = int(numero_semana), int(anio)
        # Crear un objeto de fecha y tiempo para el primer día del año dado
        primer_dia_anio = datetime.datetime(anio, 1, 1)

        # Calcular la diferencia en días entre el primer día del año y el primer día de la semana dada
        dias_diferencia = (numero_semana - 1) * 7

        # Obtener el primer día de la semana sumando la diferencia de días al primer día del año
        primer_dia_semana = primer_dia_anio + datetime.timedelta(days=dias_diferencia)

        return primer_dia_semana
    
class ExcelFunctions:
    def write_multiple_cells(self, working_sheet, starting_row=None, starting_col=None, values=None):
        '''
        Esta función escribe todos los valores que se pasen en values sobre la misma fila.
        Empieza en starting_row, y termina hasta quedarse sin valores para escribir
        sheet_name es el nombre de la hoja a escribir
        '''
        for index, value in enumerate(values):
            working_sheet.cell(starting_row, starting_col+index).value = value

    def find_value(self, working_sheet, col, text, initial_row=1, max_consecutive_nones=10):
        """
        Función que regresa las filas en las que se encuentra un valor específico (text)
        Lo regresa en formato lista.
        hoja: la hoja sobre la que se está trabajando
        col: número de la columna sobre la que se desea iterar
        initial_row: la primera fila a partir de la cual se desea empezar a buscar
        text: el valor específico que se busca en la columna
        max_consecutive_nones: número máximo de nones antes de parar
        """
        # Si nos dan la letra de la columna la convertimos a número para poder trabajar
        if type(col) is str:
            col = coordinate_to_tuple(f'{col}1')[1]
            
        found_cells = []
        row = initial_row
        consecutive_nones = 0
        while consecutive_nones < max_consecutive_nones:
            value = working_sheet.cell(row, col).value
            # Hay dos opciones: value es el valor que buscamos o no
            if value == text:
                found_cells.append(row)
            else: # si no es el valor que buscamos tenemos que ver si es None o no
                if value is None: # Si el valor es None, tenemos que agregar 1 al contador de Nones
                    consecutive_nones += 1
                else: # Si el valor no es None, regresamos a 0 el contador de consecutive Nones
                    consecutive_nones = 0
            row += 1
        return found_cells

    def get_columns_on_row(self, working_sheet, row, num_nones_consecutivos=3, starting_col=1):
        '''
        Función que regresa en un diccionario los nombres de las columnas con sus correspondientes columnas 
        en número. Se detiene tras haber encontrado num_nones_consecutivos nones consecutivos.
        hoja: working sheet
        numero fila: fila sobre la que se itera
        num_nones_consecutivos: número de nones consecutivos para detenerse
        starting_col: columna en la que se iniciará (default es columna A)
        '''
        diccionario_columnas = {}
        nones_consecutivos = 0
        
        for columna, celda in enumerate(working_sheet[row], start=starting_col):
            valor_celda = celda.value
            
            if valor_celda is not None:
                diccionario_columnas[valor_celda] = columna
                nones_consecutivos = 0
            else:
                nones_consecutivos += 1
                if nones_consecutivos >= num_nones_consecutivos:
                    break
        
        return diccionario_columnas

    def find_constrained_data(self, working_sheet, start_row, start_column):
        """
        Esta función encuentra toda la información contenida entre Nones de forma horizontal
        y vertical a partir de una celda dada. Regresa un numpy array con los datos.
        
        ws: working sheet, esto es una hoja de excel de openpyxl
        start_row, satrt_column: las coordenadas de la celda en la que se empieza a buscar los datos
        
        Warning: para que esta función sea óptima, es necesario que no haya Nones en todos los datos 
        sobre los cuales se está iterando, de otra forma se parará de forma temprana
        """
        data = []
        
        if type(start_column) is str:
            column = coordinate_to_tuple(f'{start_column}1')[1]
        else:
            column = start_column
        while True:
            row = start_row
            column_data = []
            
            cell = working_sheet.cell(row=row, column=column)
            while cell.value is not None:
                column_data.append(cell.value)
                row += 1
                cell = working_sheet.cell(row=row, column=column)
            
            if not column_data:
                break
            
            data.append(column_data)
            column += 1
        
        return np.array(data).transpose()

    def get_last_col_with_data(self, working_sheet, row, starting_col=1, step=1, end_indicator=None):
        """
        Esta función recibe una hoja de trabajo, una fila, una columna en la que se inician
        los datos y un salto entre columnas con datos
        
        Itera de forma horizontal cada 'step' columnas hasta que alguna sea None para saber
        cuándo deja de haber datos
        
        Regresa un entero con el valor numérico de la última columna con datos + step - 1
        """
        actual = starting_col
        previo = -1
        
        # Ahora tenemos que ver si end_indicator es None o no
        if end_indicator is None:
            while(1):
                if working_sheet.cell(row, actual).value is not None:
                    # Si no es None, puede ser también de la forma MES DE ABRIL 2020
                    contenido_celda = working_sheet.cell(row, actual).value
                    primera_palabra = contenido_celda.split()[0]
                    if primera_palabra != "MES":
                        previo = actual
                        actual += step
                    else:
                        break
                else:
                    break
        else:
            while(1):
                if working_sheet.cell(row, actual).value != end_indicator:
                    previo = actual
                    actual += step
                else:
                    break
        return previo + step - 1

    def get_last_row_with_data(self, working_sheet, column, starting_row, text_at_end, padding_between_end=0, max_row=100_000):
        """
        Esta función recibe una hoja de trabajo, una columna, una fila en la que se inician
        los datos y un posible padding entre el dato que indica el final y el final real de
        los datos que contienen la información necesaria
        
        Itera de forma vertical empezando en 'starting_row' hasta que encuentre 'text_at_end'
        
        Regresa un entero con el valor numérico de la última fila con datos menos el padding
        especificado
        """
        actual = starting_row
        previo = -1
        
        while(1):
            if working_sheet.cell(actual, column).value != text_at_end:
                previo = actual
                actual += 1
            else:
                if actual > max_row:
                    return -1
                break
                
        return previo - padding_between_end

    def get_data_on_columns(self, working_sheet, starting_row, dict_cols, ending_row=None, cols_to_extract=None, output_columns=None, rows_to_extract=100):
        '''
        Esta función extrae los datos de una hoja de excel y lo regresa como un dataframe.
        working_sheet: la hoja de excel
        starting_row: el número de la fila en la que empiezan los valores
        dict_cols: un diccionario de la siguiente forma {'col_name_1':numero_columna_1, 'col_name_2':numero_columna_2}
        ending_row: el número de la fila final donde terminan los valores (si es None, se extraen rows_to_extract filas tras starting_row)
        cols_to_extract: una lista con los nombres de las columnas que se desean seleccionar (None->todas las dict_cols)
        output_columns: los nombres de las columnas que se usarán en el DataFrame (None-> las mismas que dict_cols)
        rows_to_extract: el máximo número de filas a extraer. Solo se considera si ending_row es None
        '''
        # Definimos los parámetros necesarios
        
        # LAST_ROW
        if ending_row is None:
            last_row = starting_row + rows_to_extract
        else:
            last_row = ending_row
            
        # COLS_TO_EXTRACT
        if cols_to_extract is None:
            cols_to_extract = list(dict_cols.keys())
        
        # OUTPUT_COLUMNS
        if output_columns is None:
            output_columns = cols_to_extract
            
        listas_datos = [[] for _ in range(len(cols_to_extract))]
        for row in np.arange(starting_row, last_row):
            for lista, col_name in zip(listas_datos, cols_to_extract):
                lista.append(working_sheet.cell(row, dict_cols[col_name]).value)
                
        df = pd.DataFrame(np.array(listas_datos).transpose(), columns=output_columns)
        df.dropna(how='all', inplace=True)
        return df 

    def encontrar_ultima_fila_con_datos(self, working_sheet, column, starting_row=1, consecutive_nones=3):
        # Inicializar variables
        fila_actual = starting_row
        nones_consecutivos = 0

        while 1:
            # Obtener el valor de la celda en la columna especificada
            valor_celda = working_sheet.cell(row=fila_actual, column=column).value

            if valor_celda is not None:
                # Si hay datos en la celda, reiniciamos el contador de Nones consecutivos
                nones_consecutivos = 0
            else:
                # Si la celda está vacía, incrementamos el contador de Nones consecutivos
                nones_consecutivos += 1

                # Si encontramos "consecutive_nones" Nones consecutivos, detenemos la búsqueda
                if nones_consecutivos == consecutive_nones:
                    break

            # Pasamos a la siguiente fila
            fila_actual += 1

        # Restamos 1 a la fila actual para obtener la última fila con datos
        ultima_fila_con_datos = fila_actual - consecutive_nones

        return ultima_fila_con_datos 
    
    def save_and_download_excel_file(self, df: pd.DataFrame, dir_location, file_name, sheet_name='hoja_procesada', index=False, n_cols_to_bold=0, return_data=True):
        '''
        Esta función recibe un data frame, lo convierte a excel y regresa los bytes listos para descargar junto
        con el archivo del nombre.

        df: el dataframe a convertir
        dir_location: la ubicación relativa del archivo
        file_name: el nombre del archivo
        sheet_name: el nombre de la hoja de excel a guardar
        index: si se desea guardar el index del df o no
        n_cols_to_bold: número de columnas que se pondrá en bold
        '''
        if df is None:
            return None, None
        n_rows, n_cols = df.shape

        # Creamos variables de estilo
        color, font_color = '023e8a', 'edf2f4'
        fill_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
        font = Font(color=font_color, bold=True, size=12)
        align = Alignment(horizontal='center')
        border = Border(right=Side(style='thin'))

        # Creamos un nuevo libro de Excel utilizando openpyxl
        wb = xl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Pasamos el DataFrame a la hoja de Excel utilizando 'dataframe_to_rows'
        for r in dataframe_to_rows(df, index=index, header=True):
            ws.append(r)

        # Cambiamos la anchura de las columnas
        for col in range(n_cols):
            col = get_column_letter(col+1)
            ws.column_dimensions[col].adjustment = .3

        # Cambiamos el color de la primera fila
        for cell in ws[1]:
            cell.fill = fill_color
            cell.font = font
            cell.alignment = align

        if n_cols_to_bold > 0 and n_cols_to_bold <= n_cols:
            # Cambiamos el tipo de fuente de las primeras dos a bold
            for row in np.arange(start=2, stop=n_rows+2):
                for col in np.arange(start=1, stop=1+n_cols_to_bold):
                    ws.cell(row=row, column=col).font = Font(bold=True)

            # Cambiamos el borde de la segunda columna
            for row in np.arange(start=2, stop=n_rows+2):
                ws.cell(row=row, column=n_cols_to_bold).border = border
                ws.cell(row=row, column=n_cols_to_bold).alignment = align    

        # Ajustamos la anchura de las columnas
        self.adjust_column_widths(ws)
        # Guardamos el archivo procesado
        wb.save(f"{dir_location}/{file_name}.xlsx")

        # Una vez guardado el archivo, lo podemos regresar como bytes
        with open(f"{dir_location}/{file_name}.xlsx", "rb") as f:
            bytes_data = f.read()

        if return_data:
            return bytes_data, f'{file_name}.xlsx'
        else:
            return None, None
        
    def get_columns_on_row_repeated_names(self, working_sheet, row, num_nones_consecutivos=3, starting_col=1):
        '''
        Función que regresa en un diccionario los nombres de las columnas con sus correspondientes columnas 
        en número. Se detiene tras haber encontrado num_nones_consecutivos nones consecutivos.
        hoja: working sheet
        numero fila: fila sobre la que se itera
        num_nones_consecutivos: número de nones consecutivos para detenerse
        starting_col: columna en la que se iniciará (default es columna A)
        '''
        diccionario_columnas = {}
        nones_consecutivos = 0
        
        for columna, celda in enumerate(working_sheet[row], start=starting_col):
            valor_celda = celda.value
            
            if valor_celda is not None:
                if valor_celda not in diccionario_columnas:
                    diccionario_columnas[valor_celda] = [columna]
                else:
                    diccionario_columnas[valor_celda].append(columna)
                nones_consecutivos = 0
            else:
                nones_consecutivos += 1
                if nones_consecutivos >= num_nones_consecutivos:
                    break
        
        return diccionario_columnas
    
    def adjust_column_widths(self, working_sheet):
        """
        Esta función recibe una hoja de excel y ajusta la anchura de todas las columnas.
        Parece que solo funciona si las columnas están en la primera fila. Creo.
        """
        for columna in working_sheet.columns:
            max_length = 0
            columna_letra = get_column_letter(columna[0].column)
            for celda in columna:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(celda.value)
                except:
                    pass
            ajuste_anchura = (max_length + 2)
            working_sheet.column_dimensions[columna_letra].width = ajuste_anchura