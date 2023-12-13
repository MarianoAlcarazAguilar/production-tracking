import pandas as pd
import numpy as np
import datetime
import os
from openpyxl.utils.cell import coordinate_to_tuple
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# La idea que tengo es la siguiente. 
# En general, la mayoría de las funciones que tengo son para extraer información
# Y el problema que tengo es que siempre le tengo que estar pasando el nombre del archivo
# o algo del estilo. Lo que haré ahora es separar las funciones por tipo.
# Unas serán de extracción de información, y las otras serán de modificación de 
# archivos existes.

class DataExtraction:
    def __init__(self, file, sheet_name=0):
        '''
        Se cargará al objeto los datos que se deseen manipular para extraer información.
        Se tendrá de dos maneras. Una en pd.DataFrame y otra en xl.WorkingBook.

        sheet_name: la hoja que se desea trabajar. Por default es la active sheet.
        '''
        self.df_file = self.__open_excel_file_as_dataframe(file, sheet_name=sheet_name)
        #self.wb_file, self.ws_file = self.__open_excel_file_as_working_book(file, sheet_name=sheet_name)

    def __open_excel_file_as_dataframe(self, file, sheet_name=0) -> pd.DataFrame:
        '''
        TODO: Probar que file puede ser una ubicación de archivo y bytes

        Esta función abre el archivo de excel especificado como un pd.DataFrame,
        pero mantiene la misma estructura de las celdas. Las filas empiezan en 1 
        y las columnas empiezan en 1. Así, la esquina superior izquierda (A1) será
        (1,1)
        '''
        df = (pd
         .read_excel(file, skiprows=0, header=None, sheet_name=sheet_name)
         .rename(lambda x: x+1)
         .rename(columns=lambda x: x+1)
        )
        return df
    
    def __open_excel_file_as_working_book(self, file, sheet_name=0) -> (xl.Workbook, xl.Workbook.worksheets):
        '''
        Esta función abre el archivo de excel especificado como un xl.Workbook. Y la hoja deseada.
        '''
        wb = xl.load_workbook(file, data_only=True)
        if sheet_name == 0:
            ws = wb.active
        else:
            # Nos aseguramos de que la hoja que nos dieron esté en el archivo
            assert sheet_name in wb.worksheets, "The given sheet does not exist"
            ws = wb[sheet_name]
        return (wb, ws)
    
    def find_value(self, value_to_find) -> list:
        '''
        Esta función se utiliza para encontrar el value_to_find en todo el excel.
        Usa pandas para que sin importar el tamaño del archivo, el tiempo que se 
        tarde en encontrarlo sea el mismo.

        Regresa una lista de tuplas con las coordenadas convertidas a celdas de excel.
        Ejemplo: si buscas la palabra 'hola' y aparece en la celda A1 y C3 regresa [(1,1), (3,3)]
        Esto es importante recalcarlo porque en pandas empieza en 0, mientras que en excel empieza en 1

        Si el valor no se encuetra, se regresa una lista vacía
        '''

        result = (
            self
            .df_file
            .eq(value_to_find)
            .stack()
            .pipe(lambda df: df[df])
            .index
            .to_list()
        )

        return result

    def get_value_on_cell(self, cell:(int, int)):
        '''
        Esta función regresa el valor de la celda especificada
        '''
        assert len(cell) == 2, "La celda necesita ser de tamaño 2"
        assert cell[0] >= 1, "La celda no está bien especificada"
        assert cell[1] >= 1, "La celda no está bien especificada"
        return self.df_file.loc[cell]

    def extract_data_from_file(self, index_value, cols_to_extract:list, shift_between_values:int=0) -> pd.DataFrame:
        '''
        Esta función es la última función que vas a necesitar.
        Solo dile cuál es el nombre de la columna que se usa como índice, el nombre de las columnas a extraer
        y el total de celdas que exista (o no) de diferencia entre la fila de los índices y donde se encuentren
        los valores.

        Inputs:
            - index_value: el valor de la celda donde están los índices. Debe estar a la izquierda (no inmediata necesariamente) de cols_to_extract
            - cols_to_extract: el nombre de las columnas que quieres extraer. Por default se extrae también index_value col
            - shift_between_values: la diferencia entre las filas donde están los índices y los valores

        Returns:
            - pd.DataFrame con los valores extraídos

        Example:
        excel_functions.extract_data_from_file('/data/file_mariano.xlsx', 'clave', ['hora_entrada', 'dia'], shift_between_values=2)
        '''
        # Paso 1
        # Encontrar las filas donde están las columnas que queremos extraer
        # Siempre extraemos el valor del index
        cols_to_extract = [index_value] + cols_to_extract
        dic_cells = {llave:self.find_value(llave)[0] for llave in cols_to_extract}

        row_index_cell, col_index_cell = dic_cells[index_value]

        # Paso 2
        # Encontrar los índices de las columnas a extraer
        cols_to_extract_index = [dic_cells[llave][1] for llave in cols_to_extract]

        dic_to_rename_columns = {llave:valor for llave, valor in zip(cols_to_extract_index, cols_to_extract)}

        # Paso 3
        # Sacar los datos
        output = (
            self
            .df_file
            .loc[row_index_cell+1:, cols_to_extract_index]
            .set_index(col_index_cell)
            .shift(shift_between_values)
            .reset_index()
            .dropna()
            .rename(dic_to_rename_columns, axis=1)
            .reset_index(drop=True)
        )

        return output
    

    def get_column_names_on_row(self, row:int) -> list:
        '''
        Esta función regresa los nombres de la columa especificada
        '''
        assert row >= 1, "la fila no es válida"
        return self.df_file.loc[row].dropna().to_list()